import logging
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from data_fetcher import get_etf_symbols, get_cfd_symbols, get_forex_symbols, get_data
from movers_calculator import get_top_movers
from analyzer import analyze_data, analyze_movement, estimate_max_profit, calculate_tp, estimate_duration
from technical_indicators import add_technical_indicators
from visualization import plot_price_and_indicators, plot_backtest_results, plot_regression_results, plot_clustering_results
from backtesting import backtest_strategy, simple_moving_average_strategy, simple_strategy, rsi_strategy, write_backtest_log
from machine_learning import train_regression_model, train_random_forest_model, predict_price, train_clustering_model, cluster_data

# Fichiers à supprimer
log_file = 'trade.log'
excel_file = 'trade.xlsx'
backtest_log_file = 'backtest_log.txt'

# Supprimer les fichiers s'ils existent
if os.path.exists(log_file):
    os.remove(log_file)
if os.path.exists(excel_file):
    os.remove(excel_file)
if os.path.exists(backtest_log_file):
    os.remove(backtest_log_file)

# Créer le répertoire de sortie pour les graphiques
output_dir = "plots"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Configuration de la journalisation
logging.basicConfig(filename=log_file, level=logging.DEBUG, 
                    format='%(asctime)s %(levelname)s: %(message)s')

def check_close_column(data, symbols):
    missing_close = []
    for symbol in symbols:
        if symbol in data and 'Close' in data[symbol].columns:
            logging.debug(f"Column 'Close' found in data for symbol: {symbol}")
        else:
            logging.error(f"Column 'Close' not found in data for symbol: {symbol}")
            missing_close.append(symbol)
    return missing_close

def calculate_risk_management(data, action, capital=10000, risk_per_trade=0.01):
    stop_loss_pct = 0.02  # 2% du prix actuel
    take_profit_pct = 0.04  # 4% du prix actuel
    current_price = data['Close'].iloc[-1]

    if action == 'buy':
        stop_loss_price = current_price * (1 - stop_loss_pct)
        take_profit_price = current_price * (1 + take_profit_pct)
    elif action == 'short':
        stop_loss_price = current_price * (1 + stop_loss_pct)
        take_profit_price = current_price * (1 - take_profit_pct)
    else:
        stop_loss_price = current_price
        take_profit_price = current_price

    # Éviter la division par zéro
    if current_price == stop_loss_price:
        position_size = 0
    else:
        position_size = (capital * risk_per_trade) / abs(current_price - stop_loss_price)

    return position_size, stop_loss_price, take_profit_price

def write_to_excel(filename, data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Data"

    for row in data:
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                if "Gainer" in cell.value:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif "Loser" in cell.value:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    wb.save(filename)

def evaluate_reliability(reg_pred, rf_pred, action, tp):
    if action == "buy" and reg_pred > tp and rf_pred > tp:
        return "High Confidence"
    elif action == "short" and reg_pred < tp and rf_pred < tp:
        return "High Confidence"
    else:
        return "Low Confidence"

def main():
    logging.info('Starting main script')
    try:
        # Récupérer les symboles et les données pour les ETF
        etf_symbols = get_etf_symbols()
        etf_data = get_data(etf_symbols, period="5d", interval="1d")
        logging.debug(f'ETF Data: {etf_data.head()}')
        top_gainers_etf, top_losers_etf = get_top_movers(etf_data)
        missing_close_etf = check_close_column(etf_data, etf_symbols)
        if missing_close_etf:
            raise ValueError(f"Column 'Close' not found in ETF data for symbols: {missing_close_etf}")

        # Récupérer les symboles et les données pour les CFD
        cfd_symbols = get_cfd_symbols()
        cfd_data = get_data(cfd_symbols, period="1d", interval="1m")
        logging.debug(f'CFD Data: {cfd_data.head()}')
        top_gainers_cfd, top_losers_cfd = get_top_movers(cfd_data)
        missing_close_cfd = check_close_column(cfd_data, cfd_symbols)
        if missing_close_cfd:
            raise ValueError(f"Column 'Close' not found in CFD data for symbols: {missing_close_cfd}")

        # Journalisation pour CFD
        logging.debug(f'Top Gainers CFDs: {top_gainers_cfd}')
        logging.debug(f'Top Losers CFDs: {top_losers_cfd}')

        # Récupérer les symboles et les données pour les paires Forex
        forex_symbols = get_forex_symbols()
        forex_data = get_data(forex_symbols, period="1mo", interval="1d")
        logging.debug(f'Forex Data: {forex_data.head()}')
        top_gainers_forex, top_losers_forex = get_top_movers(forex_data)
        missing_close_forex = check_close_column(forex_data, forex_symbols)
        if missing_close_forex:
            raise ValueError(f"Column 'Close' not found in Forex data for symbols: {missing_close_forex}")

        # Analyser les données et ajouter les indicateurs techniques
        logging.info("Analyzing data and adding technical indicators")
        results = analyze_data(top_gainers_etf, top_losers_etf, etf_data, top_gainers_cfd, top_losers_cfd, cfd_data, top_gainers_forex, top_losers_forex, forex_data)
        
        # Préparer les résultats au format Excel
        excel_lines = [["Category", "Symbol", "Percent Change", "Action", "TP", "Max Profit", "Duration", "Predicted Price (Reg)", "Predicted Price (RF)", "Stop Loss", "Take Profit", "Comments", "Reliability"]]
        excel_lines += [["Top Gainers ETFs (5 Days)"]]
        for item in top_gainers_etf:
            symbol, change = item
            action = analyze_movement(float(change))
            max_profit = estimate_max_profit(symbol, etf_data)
            tp = calculate_tp(symbol, etf_data, action)
            duration = estimate_duration(symbol, etf_data, action)
            
            # Prédiction du prix avec les modèles de régression
            reg_model, reg_mse, X_test, y_test, reg_pred = train_regression_model(etf_data[symbol], 'Close')
            rf_model, rf_mse, X_test, y_test, rf_pred = train_random_forest_model(etf_data[symbol], 'Close')
            reg_pred_price = reg_pred[-1] if reg_pred is not None else "N/A"
            rf_pred_price = rf_pred[-1] if rf_pred is not None else "N/A"
            
            # Calcul du stop loss et du take profit
            position_size, stop_loss_price, take_profit_price = calculate_risk_management(etf_data[symbol], action)
            
            comments = "Consistent" if action == "buy" and reg_pred_price > tp and rf_pred_price > tp else "Contradictory"
            reliability = evaluate_reliability(reg_pred_price, rf_pred_price, action, tp)
            
            excel_lines.append(["ETF Gainer", symbol, f"{change:.2f}", action, f"{tp:.2f}", f"{max_profit:.2f}", f"{duration:.2f}", f"{reg_pred_price}", f"{rf_pred_price}", f"{stop_loss_price:.2f}", f"{take_profit_price:.2f}", comments, reliability])

        excel_lines += [["Top Losers ETFs (5 Days)"]]
        for item in top_losers_etf:
            symbol, change = item
            action = analyze_movement(float(change))
            max_profit = estimate_max_profit(symbol, etf_data)
            tp = calculate_tp(symbol, etf_data, action)
            duration = estimate_duration(symbol, etf_data, action)
            
            # Prédiction du prix avec les modèles de régression
            reg_model, reg_mse, X_test, y_test, reg_pred = train_regression_model(etf_data[symbol], 'Close')
            rf_model, rf_mse, X_test, y_test, rf_pred = train_random_forest_model(etf_data[symbol], 'Close')
            reg_pred_price = reg_pred[-1] if reg_pred is not None else "N/A"
            rf_pred_price = rf_pred[-1] if rf_pred is not None else "N/A"
            
            # Calcul du stop loss et du take profit
            position_size, stop_loss_price, take_profit_price = calculate_risk_management(etf_data[symbol], action)
            
            comments = "Consistent" if action == "short" and reg_pred_price < tp and rf_pred_price < tp else "Contradictory"
            reliability = evaluate_reliability(reg_pred_price, rf_pred_price, action, tp)
            
            excel_lines.append(["ETF Loser", symbol, f"{change:.2f}", action, f"{tp:.2f}", f"{max_profit:.2f}", f"{duration:.2f}", f"{reg_pred_price}", f"{rf_pred_price}", f"{stop_loss_price:.2f}", f"{take_profit_price:.2f}", comments, reliability])

        excel_lines += [["Top Gainers CFDs (Last Day)"]]
        for item in top_gainers_cfd:
            symbol, change = item
            action = analyze_movement(float(change))
            max_profit = estimate_max_profit(symbol, cfd_data)
            tp = calculate_tp(symbol, cfd_data, action)
            duration = estimate_duration(symbol, cfd_data, action)
            
            # Prédiction du prix avec les modèles de régression
            reg_model, reg_mse, X_test, y_test, reg_pred = train_regression_model(cfd_data[symbol], 'Close')
            rf_model, rf_mse, X_test, y_test, rf_pred = train_random_forest_model(cfd_data[symbol], 'Close')
            reg_pred_price = reg_pred[-1] if reg_pred is not None else "N/A"
            rf_pred_price = rf_pred[-1] if rf_pred is not None else "N/A"
            
            # Calcul du stop loss et du take profit
            position_size, stop_loss_price, take_profit_price = calculate_risk_management(cfd_data[symbol], action)
            
            comments = "Consistent" if action == "buy" and reg_pred_price > tp and rf_pred_price > tp else "Contradictory"
            reliability = evaluate_reliability(reg_pred_price, rf_pred_price, action, tp)
            
            excel_lines.append(["CFD Gainer", symbol, f"{change:.2f}", action, f"{tp:.2f}", f"{max_profit:.2f}", f"{duration:.2f}", f"{reg_pred_price}", f"{rf_pred_price}", f"{stop_loss_price:.2f}", f"{take_profit_price:.2f}", comments, reliability])

        excel_lines += [["Top Losers CFDs (Last Day)"]]
        for item in top_losers_cfd:
            symbol, change = item
            action = analyze_movement(float(change))
            max_profit = estimate_max_profit(symbol, cfd_data)
            tp = calculate_tp(symbol, cfd_data, action)
            duration = estimate_duration(symbol, cfd_data, action)
            
            # Prédiction du prix avec les modèles de régression
            reg_model, reg_mse, X_test, y_test, reg_pred = train_regression_model(cfd_data[symbol], 'Close')
            rf_model, rf_mse, X_test, y_test, rf_pred = train_random_forest_model(cfd_data[symbol], 'Close')
            reg_pred_price = reg_pred[-1] if reg_pred is not None else "N/A"
            rf_pred_price = rf_pred[-1] if rf_pred is not None else "N/A"
            
            # Calcul du stop loss et du take profit
            position_size, stop_loss_price, take_profit_price = calculate_risk_management(cfd_data[symbol], action)
            
            comments = "Consistent" if action == "short" and reg_pred_price < tp and rf_pred_price < tp else "Contradictory"
            reliability = evaluate_reliability(reg_pred_price, rf_pred_price, action, tp)
            
            excel_lines.append(["CFD Loser", symbol, f"{change:.2f}", action, f"{tp:.2f}", f"{max_profit:.2f}", f"{duration:.2f}", f"{reg_pred_price}", f"{rf_pred_price}", f"{stop_loss_price:.2f}", f"{take_profit_price:.2f}", comments, reliability])

        excel_lines += [["Top Gainers Forex Pairs (Last Day)"]]
        for item in top_gainers_forex:
            symbol, change = item
            action = analyze_movement(float(change))
            max_profit = estimate_max_profit(symbol, forex_data)
            tp = calculate_tp(symbol, forex_data, action)
            duration = estimate_duration(symbol, forex_data, action)
            
            # Prédiction du prix avec les modèles de régression
            reg_model, reg_mse, X_test, y_test, reg_pred = train_regression_model(forex_data[symbol], 'Close')
            rf_model, rf_mse, X_test, y_test, rf_pred = train_random_forest_model(forex_data[symbol], 'Close')
            reg_pred_price = reg_pred[-1] if reg_pred is not None else "N/A"
            rf_pred_price = rf_pred[-1] if rf_pred is not None else "N/A"
            
            # Calcul du stop loss et du take profit
            position_size, stop_loss_price, take_profit_price = calculate_risk_management(forex_data[symbol], action)
            
            comments = "Consistent" if action == "buy" and reg_pred_price > tp and rf_pred_price > tp else "Contradictory"
            reliability = evaluate_reliability(reg_pred_price, rf_pred_price, action, tp)
            
            excel_lines.append(["Forex Gainer", symbol, f"{change:.2f}", action, f"{tp:.2f}", f"{max_profit:.2f}", f"{duration:.2f}", f"{reg_pred_price}", f"{rf_pred_price}", f"{stop_loss_price:.2f}", f"{take_profit_price:.2f}", comments, reliability])

        excel_lines += [["Top Losers Forex Pairs (Last Day)"]]
        for item in top_losers_forex:
            symbol, change = item
            action = analyze_movement(float(change))
            max_profit = estimate_max_profit(symbol, forex_data)
            tp = calculate_tp(symbol, forex_data, action)
            duration = estimate_duration(symbol, forex_data, action)
            
            # Prédiction du prix avec les modèles de régression
            reg_model, reg_mse, X_test, y_test, reg_pred = train_regression_model(forex_data[symbol], 'Close')
            rf_model, rf_mse, X_test, y_test, rf_pred = train_random_forest_model(forex_data[symbol], 'Close')
            reg_pred_price = reg_pred[-1] if reg_pred is not None else "N/A"
            rf_pred_price = rf_pred[-1] if rf_pred is not None else "N/A"
            
            # Calcul du stop loss et du take profit
            position_size, stop_loss_price, take_profit_price = calculate_risk_management(forex_data[symbol], action)
            
            comments = "Consistent" if action == "short" and reg_pred_price < tp and rf_pred_price < tp else "Contradictory"
            reliability = evaluate_reliability(reg_pred_price, rf_pred_price, action, tp)
            
            excel_lines.append(["Forex Loser", symbol, f"{change:.2f}", action, f"{tp:.2f}", f"{max_profit:.2f}", f"{duration:.2f}", f"{reg_pred_price}", f"{rf_pred_price}", f"{stop_loss_price:.2f}", f"{take_profit_price:.2f}", comments, reliability])

        # Écrire les résultats dans le fichier Excel
        write_to_excel(excel_file, excel_lines)

        # Visualisation des données
        for item in top_gainers_etf + top_losers_etf:
            symbol, _ = item
            logging.info(f"Plotting price and indicators for ETF: {symbol}")
            if symbol in etf_data and 'Close' in etf_data[symbol].columns:
                etf_indicators = add_technical_indicators(etf_data[symbol])
                plot_price_and_indicators(symbol, etf_data[symbol], etf_indicators, output_dir)
            else:
                logging.error(f"Column 'Close' not found in ETF data for symbol: {symbol}")

        for item in top_gainers_cfd + top_losers_cfd:
            symbol, _ = item
            logging.info(f"Plotting price and indicators for CFD: {symbol}")
            if symbol in cfd_data and 'Close' in cfd_data[symbol].columns:
                cfd_indicators = add_technical_indicators(cfd_data[symbol])
                plot_price_and_indicators(symbol, cfd_data[symbol], cfd_indicators, output_dir)
            else:
                logging.error(f"Column 'Close' not found in CFD data for symbol: {symbol}")

        for item in top_gainers_forex + top_losers_forex:
            symbol, _ = item
            logging.info(f"Plotting price and indicators for Forex: {symbol}")
            if symbol in forex_data and 'Close' in forex_data[symbol].columns:
                forex_indicators = add_technical_indicators(forex_data[symbol])
                plot_price_and_indicators(symbol, forex_data[symbol], forex_indicators, output_dir)
            else:
                logging.error(f"Column 'Close' not found in Forex data for symbol: {symbol}")

        # Backtesting des stratégies de trading
        for symbol in etf_symbols:
            logging.debug(f"Data for {symbol}: {etf_data[symbol].head()}")
            strategy = simple_moving_average_strategy(etf_data[symbol])
            if strategy.empty:
                logging.error(f"Strategy for {symbol} could not be computed.")
                continue
            final_value, trading_log = backtest_strategy(etf_data[symbol], symbol, strategy)
            logging.info(f"Backtesting {symbol}: Final portfolio value: {final_value}")
            write_backtest_log(symbol, trading_log, backtest_log_file)
            plot_backtest_results(symbol, etf_data[symbol], trading_log, output_dir)

            # Test with simple strategy
            simple_strat = simple_strategy(etf_data[symbol])
            if simple_strat.empty:
                logging.error(f"Simple strategy for {symbol} could not be computed.")
                continue
            final_value, trading_log = backtest_strategy(etf_data[symbol], symbol, simple_strat)
            logging.info(f"Backtesting with simple strategy {symbol}: Final portfolio value: {final_value}")
            write_backtest_log(f"{symbol}_simple_strategy", trading_log, backtest_log_file)
            plot_backtest_results(f"{symbol}_simple_strategy", etf_data[symbol], trading_log, output_dir)

            # Test with RSI strategy
            rsi_strat = rsi_strategy(etf_data[symbol])
            if rsi_strat.empty:
                logging.error(f"RSI strategy for {symbol} could not be computed.")
                continue
            final_value, trading_log = backtest_strategy(etf_data[symbol], symbol, rsi_strat)
            logging.info(f"Backtesting with RSI strategy {symbol}: Final portfolio value: {final_value}")
            write_backtest_log(f"{symbol}_rsi_strategy", trading_log, backtest_log_file)
            plot_backtest_results(f"{symbol}_rsi_strategy", etf_data[symbol], trading_log, output_dir)

            # Train regression model and plot results
            clean_data = etf_data[symbol].dropna()
            if not clean_data.empty:
                model, mse, X_test, y_test, y_pred = train_regression_model(clean_data, 'Close')
                if model:
                    logging.info(f"Trained regression model for {symbol} with MSE: {mse}")
                    plot_regression_results(X_test, y_test, y_pred, output_dir)

            # Train clustering model and plot results
            clustering_data = etf_data[symbol][['Close', 'Volume']].dropna()
            if not clustering_data.empty:
                clusters_model, silhouette_avg, clusters = train_clustering_model(clustering_data, n_clusters=3)
                if clusters_model:
                    logging.info(f"Trained clustering model for {symbol} with silhouette score: {silhouette_avg}")
                    plot_clustering_results(clustering_data, clusters, output_dir)

        logging.info('Script completed successfully')
    except Exception as e:
        logging.error(f'Error in main script: {e}', exc_info=True)

if __name__ == "__main__":
    main()

print("Script completed successfully")
